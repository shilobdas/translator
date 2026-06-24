from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from .internal_config import internal_translate_text
from .translation import TranslationServiceError


XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass
class TranslatedExcel:
    filename: str
    content: bytes
    mime_type: str
    target_languages: list[str]
    character_count: int
    rows_translated: int


def translate_xlsx_bytes(
    filename: str,
    content: bytes,
    source_language: str,
    target_languages: list[str],
    columns: list[str],
) -> TranslatedExcel:
    if Path(filename).suffix.lower() != ".xlsx":
        raise TranslationServiceError("Unsupported spreadsheet type. Upload .xlsx.")
    selected_columns = [column.strip() for column in columns if column.strip()]
    if not selected_columns:
        raise TranslationServiceError("Select at least one column to translate.")

    workbook = load_workbook(BytesIO(content))
    character_count = 0
    translated_rows = set()
    matched_column_count = 0

    for worksheet in workbook.worksheets:
        headers = _header_map(worksheet)
        selected_indexes = {
            column: headers[column]
            for column in selected_columns
            if column in headers
        }
        if not selected_indexes:
            continue
        matched_column_count += len(selected_indexes)

        output_indexes = {}
        next_column = worksheet.max_column + 1
        for column_name in selected_indexes:
            for target_language in target_languages:
                output_header = f"{column_name}_{target_language}"
                worksheet.cell(row=1, column=next_column, value=output_header)
                output_indexes[(column_name, target_language)] = next_column
                next_column += 1

        for row_index in range(2, worksheet.max_row + 1):
            for column_name, source_column_index in selected_indexes.items():
                value = worksheet.cell(row=row_index, column=source_column_index).value
                if value is None:
                    continue
                source_text = str(value)
                if not source_text.strip() or source_text.startswith("="):
                    continue
                translated_rows.add((worksheet.title, row_index))
                for target_language in target_languages:
                    result = internal_translate_text(
                        source_text,
                        source_language,
                        target_language,
                    )
                    output_column_index = output_indexes[(column_name, target_language)]
                    worksheet.cell(row=row_index, column=output_column_index, value=result.text)
                    character_count += len(source_text)

    if matched_column_count == 0:
        raise TranslationServiceError(
            f"Selected columns were not found in the first row: {', '.join(selected_columns)}"
        )

    output = BytesIO()
    workbook.save(output)
    return TranslatedExcel(
        filename=_translated_filename(filename),
        content=output.getvalue(),
        mime_type=XLSX_MIME_TYPE,
        target_languages=target_languages,
        character_count=character_count,
        rows_translated=len(translated_rows),
    )


def extract_xlsx_columns(content: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    columns = []
    for worksheet in workbook.worksheets:
        for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), []):
            if value is not None and str(value).strip():
                columns.append(str(value).strip())
        break
    workbook.close()
    return columns


def _header_map(worksheet) -> dict[str, int]:
    headers = {}
    for cell in worksheet[1]:
        if cell.value is None:
            continue
        header = str(cell.value).strip()
        if header and header not in headers:
            headers[header] = cell.column
    return headers


def _translated_filename(filename: str) -> str:
    path = Path(filename)
    return f"{path.stem}_translated{path.suffix}"
