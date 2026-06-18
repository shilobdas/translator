import base64
from io import BytesIO
import os
from pathlib import Path
import sys
import tempfile

os.environ["ALLOW_PUBLIC_REGISTRATION"] = "false"
os.environ["INTERNAL_TRANSLATION_PROVIDER"] = "nllb"
os.environ["NLLB_TRANSLATION_URL"] = "http://nllb.test/translate"
os.environ["SECRET_KEY"] = "smoke-secret-key"
SMOKE_DB_PATH = Path(tempfile.gettempdir()) / "translator_app_internal_smoke.db"
if SMOKE_DB_PATH.exists():
    SMOKE_DB_PATH.unlink()
os.environ["DATABASE_URL"] = f"sqlite:///{SMOKE_DB_PATH}"

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from docx import Document
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from backend.admin_tools import create_or_update_admin
from backend.main import app
from backend.services import translation as translation_service


class FakeResponse:
    status_code = 200
    text = '{"translated_text": "..."}'

    def __init__(self, translated_text):
        self._translated_text = translated_text

    def json(self):
        return {"translated_text": self._translated_text}


def fake_post(url, **kwargs):
    payload = kwargs.get("json", {})
    text = payload.get("text", "")
    target = payload.get("target_language") or payload.get("target") or "xx"
    return FakeResponse(f"{text} [{target}]")


translation_service.requests.post = fake_post
client = TestClient(app)


def login(username, password):
    response = client.post(
        "/api/auth/login",
        data={"username": username, "password": password},
    )
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_internal_smoke():
    create_or_update_admin("smoke_admin", "SmokePass123!")
    admin_headers = login("smoke_admin", "SmokePass123!")

    created = client.post(
        "/api/internal/admin/users",
        headers=admin_headers,
        json={
            "username": "smoke_user",
            "password": "SmokeUser123!",
            "is_active": True,
            "is_admin": False,
        },
    )
    assert created.status_code in {200, 400}, created.text

    user_headers = login("smoke_user", "SmokeUser123!")

    health = client.get("/api/internal/provider/health", headers=user_headers)
    assert health.status_code == 200, health.text
    assert health.json()["ok"] is True

    text_translation = client.post(
        "/api/internal/translate/text",
        headers=user_headers,
        json={
            "text": "Hello internal user",
            "source_language": "en",
            "target_language": "bn",
        },
    )
    assert text_translation.status_code == 200, text_translation.text
    assert text_translation.json()["translated_text"].endswith("[bn]")

    doc = Document()
    doc.add_paragraph("Document hello")
    doc_buffer = BytesIO()
    doc.save(doc_buffer)
    doc_response = client.post(
        "/api/internal/translate/document",
        headers=user_headers,
        data={"source_language": "en", "target_languages": "bn"},
        files={
            "file": (
                "sample.docx",
                doc_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
        },
    )
    assert doc_response.status_code == 200, doc_response.text
    doc_files = doc_response.json()["files"]
    assert doc_files[0]["filename"] == "sample_bn.docx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["description", "sku"])
    sheet.append(["Excel hello", "A1"])
    sheet.append(["Another row", "A2"])
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)

    excel_response = client.post(
        "/api/internal/translate/excel",
        headers=user_headers,
        data={
            "source_language": "en",
            "target_languages": "bn,es",
            "columns": "description",
        },
        files={
            "file": (
                "products.xlsx",
                excel_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert excel_response.status_code == 200, excel_response.text
    payload = excel_response.json()
    assert payload["rows_translated"] == 2
    translated_bytes = base64.b64decode(payload["file"]["content_base64"])
    translated_workbook = load_workbook(BytesIO(translated_bytes))
    headers = [cell.value for cell in translated_workbook.active[1]]
    assert "description_bn" in headers
    assert "description_es" in headers

    usage = client.get("/api/internal/admin/usage/summary", headers=admin_headers)
    assert usage.status_code == 200, usage.text
    totals = usage.json()["totals"]
    assert totals["text_translations"] >= 1
    assert totals["documents_uploaded"] >= 1
    assert totals["excel_files_uploaded"] >= 1
    assert totals["excel_rows_translated"] >= 2


if __name__ == "__main__":
    test_internal_smoke()
    print("internal smoke tests passed")
